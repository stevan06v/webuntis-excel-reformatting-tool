from tkinter import filedialog
import openpyxl
import numpy as np
import re
import datetime

class Assignment:

    def __init__(self, _subject, _day, _month, _type):
        self.subject = _subject
        self.type = _type
        self.date = self.create_date_of_month_day(_day, _month)

    def create_date_of_month_day(self, _day, _month):
        if _day is None or _month is None:
            raise ValueError("Invalid day or month value")

        current_year = int(datetime.datetime.now().year)

        month_number = datetime.datetime.strptime(_month, "%B").month

        return datetime.datetime(year=current_year, month=month_number, day=_day).strftime('%d/%m/%Y')

    def __str__(self):
        return f"Assignment: Subject='{self.subject}', Type='{self.type}', Date='{self.date}'"


# dictionary
dictionary = {
    "EBA_E1": "Englisch",
    "0WRRE": "Recht",
    "0WRWI": "Wirtschaft",
    "0MEDTMC": "Mobile Computing",
    "0D": "Deutsch",
    "0AM": "Mathe",
    "1SEW": "Softwareentwicklung",
    "0MEDTSM": "Social Media",
    "1MEDT3D": "Medientechnik 3D",
    "0GGPG": "Geschichte",
    "0GGPGW": "Geografie",
    "1ITP": "ITP",
    "1MEDTFI": "Medientechnik Film",
    "0NWP": "Physik",
    "1INSY": "Datenbanken",
}

german_to_english_month = {
    "Januar": "January",
    "Februar": "February",
    "März": "March",
    "April": "April",
    "Mai": "May",
    "Juni": "June",
    "Juli": "July",
    "August": "August",
    "September": "September",
    "Oktober": "October",
    "November": "November",
    "Dezember": "December"
}

# get all dictionary keys
dictionary_keys = list(dictionary.keys())

available_months = list()

# TODO: Change later on to -> ""
file_path = ""

while file_path == "":
    print("No file selected.\nRestarting...")
    # start up dialog
    file_path = filedialog.askopenfilename(
        title="Open target XLSX file",
        filetypes=(("Excel Files", "*.xlsx"),))

wb = openpyxl.load_workbook(file_path)

print("You have successfully selected: " + file_path)

ws = wb['Sheet1']

# final values
starting_date_cell_column = 2
ending_date_cell_column = 11
date_row = 2

# all list
assignment_list = list()

# loop through excel-sheet
for i in range(2, ending_date_cell_column):
    for j in range(3, 33):

        # performance fix
        current_cell_value = ws.cell(row=j, column=i).value

        if current_cell_value is not None:

            # init
            _type = ""
            _subject = ""

            month_range = j - 2
            day_range = i - 1

            _day = ws.cell(row=j, column=(i-day_range)).value
            _month = german_to_english_month.get(ws.cell(row=(j - month_range), column=i).value, None)

            # get type of assignment out of list
            for iterator in dictionary_keys:
                if iterator in current_cell_value:
                    _subject = dictionary[iterator]
                    break

            # get assignment-type
            if "Test" in current_cell_value:
                _type = "Test"
            else:
                _type = "Schularbeit"

            assignment_list.append(Assignment(_subject, _day, _month, _type))

# print the selected data
for iterator in assignment_list:
    print(iterator.__str__())


# Create a new workbook
workbook = openpyxl.Workbook()

# Select the active sheet (the default is "Sheet 0")
sheet = workbook.active

excel_headings = [
    "Fach",
    "Info",
    "Datum",
]
excel_columns = [
    "A",
    "B",
    "C",
]

# create headlines
for i in range(0, len(excel_headings)):

    sheet[excel_columns[i] + str(1)] = excel_headings[i]

    for j in range(0, len(assignment_list)):
        if i == 0:
            sheet[excel_columns[i] + str(j+2)] = str(assignment_list[j].subject)
        elif i == 1:
            sheet[excel_columns[i] + str(j+2)] = str(assignment_list[j].type)
        elif i == 2:
            sheet[excel_columns[i] + str(j+2)] = str(assignment_list[j].date)

save_file_path = ""

while save_file_path == "":
    print("No file selected.\nRestarting...")
    # start up dialog
    save_file_path = filedialog.askopenfilename(
        title="Open XLSX-file you want to save the formatted data to",
        filetypes=(("Excel Files", "*.xlsx"), ))

# Save the workbook to a file
workbook.save(save_file_path)

# Close the workbook (optional, but recommended)
workbook.close()
